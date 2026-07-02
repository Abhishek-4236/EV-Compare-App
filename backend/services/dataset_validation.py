from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import BinaryIO
from zipfile import BadZipFile

from openpyxl import load_workbook


REQUIRED_DATASET_COLUMNS = {
    "category",
    "vehicle_type",
    "brand",
    "approx_price_inr",
    "range_km",
    "battery_kwh",
}


@dataclass(frozen=True)
class DatasetValidationResult:
    rows: int
    columns: list[str]


def normalize_dataset_column(column: object) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(column).lower().strip()).strip("_")


def validate_excel_upload(
    *,
    filename: str | None,
    file_obj: BinaryIO,
    max_bytes: int,
) -> DatasetValidationResult:
    suffix = Path(filename or "").suffix.lower()
    if suffix != ".xlsx":
        raise ValueError("Only .xlsx files are allowed")

    file_obj.seek(0, 2)
    size = file_obj.tell()
    file_obj.seek(0)
    if size <= 0:
        raise ValueError("Uploaded dataset is empty")
    if size > max_bytes:
        raise ValueError(f"Uploaded dataset exceeds the {max_bytes} byte limit")

    try:
        workbook = load_workbook(file_obj, read_only=True, data_only=True)
    except BadZipFile as exc:
        raise ValueError("Uploaded file is not a valid .xlsx workbook") from exc
    except Exception as exc:
        raise ValueError("Uploaded workbook could not be opened") from exc
    finally:
        file_obj.seek(0)

    try:
        worksheet = workbook.active
        if worksheet.max_row < 2:
            raise ValueError("Dataset must contain a header row and at least one data row")

        raw_headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
        normalized_headers = [normalize_dataset_column(header) for header in raw_headers if header is not None]
        missing = sorted(REQUIRED_DATASET_COLUMNS.difference(normalized_headers))
        if missing:
            raise ValueError(f"Dataset is missing required columns: {', '.join(missing)}")

        return DatasetValidationResult(
            rows=max(0, worksheet.max_row - 1),
            columns=normalized_headers,
        )
    finally:
        workbook.close()
        file_obj.seek(0)
